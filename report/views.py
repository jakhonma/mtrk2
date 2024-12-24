from rest_framework import views, generics, mixins
from report.models import Report
from report.serializers import ReportSerializer
from rest_framework import filters, pagination
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .utils import report_info
from openpyxl.styles import Alignment


class ReportAPIView(generics.GenericAPIView):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer

    def get(self, request, *args, **kwargs):
        return self.list(request, *args, **kwargs)


class ReportListAPIView(mixins.ListModelMixin, ReportAPIView):
    pagination_class = pagination.LimitOffsetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['info']
    filterset_fields = [
        'fond__name',
        'employee__username'
    ]


class ExcelExportReportListAPIView(mixins.ListModelMixin, ReportAPIView):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['info']
    filterset_fields = [
        'fond__name',
        'employee__username'
    ]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        # Excel faylni yaratish
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "report"

        # Sarlavha qo'shish
        headers = ['Id', 'Info', 'fond', 'send', 'received', 'employee', 'number', 'dvd_number', 'name']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = header

        # Ma'lumotlarni qo'shish
        for row_num, report in enumerate(queryset, 2):
            sheet[f"A{row_num}"] = report.id
            sheet[f"B{row_num}"] = report.info
            sheet[f"C{row_num}"] = report.fond.name
            sheet[f"D{row_num}"] = report.send_mtv.name
            sheet[f"E{row_num}"] = report.received_mtv.name
            sheet[f"F{row_num}"] = report.employee.username if report.employee else ''
            sheet[f"G{row_num}"] = report.number
            sheet[f"H{row_num}"] = report.dvd_number
            sheet[f"I{row_num}"] = report_info(report)
            sheet[f"I{row_num}"].alignment = Alignment(indent=3, wrap_text=True)

        # Javobni tayyorlash
        responses = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        responses['Content-Disposition'] = 'attachment; filename="report.xlsx"'
        workbook.save(responses)
        return responses


class ReportCreateAPIView(mixins.CreateModelMixin, ReportAPIView):
    def post(self, request, *args, **kwargs):
        return self.create(request, *args, **kwargs)


class ReportUpdateAPIView(mixins.UpdateModelMixin, ReportAPIView):
    pass


class ReportDeleteAPIView(mixins.DestroyModelMixin, ReportAPIView):
    pass
